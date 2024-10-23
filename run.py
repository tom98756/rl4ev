from cmath import phase

import openpyxl

from evrp.model import Model
from evrp.env import EVREnv

import torch

import pandas as pd

from tensordict.tensordict import TensorDict

import time

start_time = time.time()



device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# Environment, Model, and Lightning Module (reinstantiate from scratch)
env = EVREnv().to(device)
model =Model(env).to(device)


# Note that by default, Lightning will call checkpoints from newer runs with "-v{version}" suffix
# unless you specify the checkpoint path explicitly
new_model_checkpoint = model.load_from_checkpoint("checkpoints/epoch_epoch=000-v13.ckpt", strict=True)
policy_new = new_model_checkpoint.policy.to(device)


# 数据处理

# 读取站台经纬度
data = pd.read_csv('data/data_50.csv', header=None)
coords = torch.tensor(data.values, dtype=torch.float32)
# 计算每个维度的最小值和最大值
min_coords = coords.min(dim=0, keepdim=True).values
max_coords = coords.max(dim=0, keepdim=True).values
# 归一化处理
normalized_coords = ((coords - min_coords) / (max_coords - min_coords)).unsqueeze(0)


# 读取站台之间的距离，单位km
data = pd.read_csv('data/distance_50.csv', header=None)
distance_matrix = torch.tensor(data.values, dtype=torch.float16)
distance_matrix.fill_diagonal_(float('inf')).unsqueeze(0)

# 读取路径是否为充电路径的数据
data = pd.read_csv('data/roads_50.csv', header=None)
charging_matrix = torch.tensor(data.values, dtype=torch.float16).unsqueeze(0)

# 读取车辆在各路径的速度
data = pd.read_csv('data/speed_50.csv', header=None)
speed_matrix = torch.tensor(data.values, dtype=torch.float16).unsqueeze(0)

charge_power = torch.tensor([100],dtype=torch.float16)


# 加载电动汽车数据
stations = pd.read_csv('data/data_50.csv', header=None, names=['lon', 'lat'])
evs_data = pd.read_csv('data/EVs_50.csv', header=None, names=['start_lon', 'start_lat', 'end_lon', 'end_lat', 'init_power','power_cap','dead_time','driving_consume'])

# 分配起始站台编号
evs_data['start_station_id'] = evs_data.apply(
    lambda row: stations[(stations['lon'] == row['start_lon']) & (stations['lat'] == row['start_lat'])].index[0], axis=1
)

# 分配目的站台编号
evs_data['end_station_id'] = evs_data.apply(
    lambda row: stations[(stations['lon'] == row['end_lon']) & (stations['lat'] == row['end_lat'])].index[0], axis=1
)

# 将rewards初始化为张量
rewards = torch.tensor([], dtype=torch.float32, device=device)

# 将起始站台编号和目的站台编号转为张量整数型
evs_data['start_station_id'] = evs_data['start_station_id'].astype(int)
evs_data['end_station_id'] = evs_data['end_station_id'].astype(int)


# 将evs_data转换为张量

# 提取需要的列
columns_to_convert = ['start_station_id', 'end_station_id', 'init_power', 'power_cap', 'dead_time', 'driving_consume']

# 创建一个字典来存储张量
evs_tensors = {}

# 遍历需要转换的列
for column in columns_to_convert:
    # 将列数据转换为张量
    if column == 'driving_consume':
        evs_tensors[column] = torch.tensor((evs_data[column].values)/100, dtype=torch.float32, device=device).unsqueeze(0)
    else:
        evs_tensors[column] = torch.tensor(evs_data[column].values, dtype=torch.float32, device=device).unsqueeze(0)
# 如果需要将所有张量合并成一个大张量，可以使用以下代码
# evs_combined_tensor = torch.stack([tensor for tensor in evs_tensors.values()], dim=1)

# 更新evs_tensors字典中的相应张量
evs_tensors['start_station_id'] = torch.tensor(evs_data['start_station_id'].values, dtype=torch.long, device=device).unsqueeze(0)
evs_tensors['end_station_id'] = torch.tensor(evs_data['end_station_id'].values, dtype=torch.long, device=device).unsqueeze(0)

# 获取电动汽车数量
elec_car_nums = evs_data.shape[0]
# 获取公交站台数量
station_nums = stations.shape[0]

available = torch.ones(
    (elec_car_nums, station_nums), dtype=torch.bool, device=device
)  # 1 means not visited, i.e. action is allowed
available[torch.arange(elec_car_nums), evs_tensors['start_station_id'].squeeze()] = False

done = torch.sum(available, dim=-1) == 0


print(f"run的规模{done.shape}")

print(distance_matrix.shape)

print(charge_power.shape)

# td = TensorDict(
#     {
#         "locs": torch.tensor(normalized_coords,dtype=torch.float32).clone().to(device),
#         "dist_mat": distance_matrix.unsqueeze(0).to(device),
#         "charge_mat": charging_matrix.to(device),
#         "car_speeds": speed_matrix.to(device),
#         "charging_ratio": charge_power.to(device),
#         "done":done.to(device)
#     },
#     batch_size=1
# )

td= TensorDict(
    {
        "locs": torch.tensor(normalized_coords,dtype=torch.float32).clone().to(device),
        "dist_mat": distance_matrix.to(device),
        "charge_mat": charging_matrix.to(device),
        "first_node": evs_tensors['start_station_id'],
        "current_node": evs_tensors['start_station_id'].unsqueeze(0),
        "prior_node": evs_tensors['start_station_id'].unsqueeze(0),
        "end_node": evs_tensors['end_station_id'].unsqueeze(0),
        "remaining_battery": evs_tensors['init_power'].unsqueeze(0),
        "bat_cap": evs_tensors['power_cap'].unsqueeze(0),
        "energy_consume_velocity": evs_tensors['driving_consume'].unsqueeze(0),
        "remaining_time": evs_tensors['dead_time'].unsqueeze(0),
        "i": torch.zeros((1, 1), dtype=torch.int64, device=device),
        "action_mask": available.unsqueeze(0),
        "stopped": torch.tensor([False], dtype=torch.bool, device=device),
        "arrived": torch.tensor([False], dtype=torch.bool, device=device),
        "done": torch.tensor([False], dtype=torch.bool, device=device),
        "reward": torch.zeros((1, 1), dtype=torch.float32, device=device)
    }
)




wb = openpyxl.Workbook()

ws1 = wb.active

ws1.title = "汇总"

ws2 = wb.create_sheet("调度详情")

rewards = torch.tensor([], dtype=torch.float32, device=device)

# out = policy_new(td.clone(), env, decode_type="greedy" , return_ations =  True)

# actions_trained = out["actions"].detach()

out = policy_new(
    td.clone(), env, phase="test", decode_type="greedy", return_actions=True
)

for i,td in enumerate(td):
    # td.update(
    #     {
    #         "first_node": evs_tensors['start_station_id'][0,i].unsqueeze(0),
    #         "current_node": evs_tensors['start_station_id'][0,i].unsqueeze(0),
    #         "prior_node": evs_tensors['start_station_id'][0,i].unsqueeze(0),
    #         "end_node": evs_tensors['end_station_id'][0,i].unsqueeze(0),
    #         "remaining_battery": evs_tensors['init_power'][0,i].unsqueeze(0),
    #         "bat_cap": evs_tensors['power_cap'][0,i].unsqueeze(0),
    #         "energy_consume_velocity": evs_tensors['driving_consume'][0,i].unsqueeze(0),
    #         "remaining_time": evs_tensors['dead_time'][0,i].unsqueeze(0),
    #         "i": torch.zeros((1, 1), dtype=torch.int64, device=device),
    #         "action_mask": available[i,:].unsqueeze(0),
    #         "stopped": torch.tensor([False], dtype=torch.bool, device=device),
    #         "arrived": torch.tensor([False], dtype=torch.bool, device=device),
    #         "done": torch.tensor([False], dtype=torch.bool, device=device),
    #         "reward": torch.zeros((1,1 ), dtype=torch.float32, device=device)
    #     }
    # )
    #with torch.inference_mode():
    # out = policy_new(
    #     td.clone(), env, phase="test", decode_type="greedy", return_actions=True
    # )

    print(f"第{i + 1}辆车的路径为{td['first_node'].item()} -> " +
          f"{' -> '.join(str(station) for station in out['actions'].squeeze().tolist())}")

    remaining_battery = td["remaining_battery"].item()

    remaining_time = td["remaining_time"].item()

    if td["arrived"]:

        remaining_battery_value = f"{remaining_battery:.2f}"

        print(f"第{i + 1}辆车的剩余电量为{remaining_battery_value}")
    else:

        remaining_battery_value = 0

        print(f"第{i+1}辆车的剩余电量为0")

    rewards = torch.cat((rewards, out["reward"].unsqueeze(0)), dim=0)

    ws1.append([i + 1, remaining_battery_value, f"{remaining_time:.2f}"])

    ws2.append([i + 1 , td["first_node"].item() , td["end_node"].item(), evs_tensors['init_power'][0,i].item(), td["bat_cap"].item() , evs_tensors['dead_time'][0,i].item()])

    actions = [td['first_node'].item()] + out['actions'].squeeze().tolist()

    path_pairs = [f"({actions[j]}, {actions[j + 1]})" for j in range(len(actions) - 1)]

    # 将所有路径对合并为一个字符串
    path_string = " ".join(path_pairs)


    # 将路径写入 ws2 表格
    ws2.append([path_string])




total_sum = sum(value[0] for value in ws1.iter_rows(min_row=1, max_col=2, min_col=2, values_only=True))

ws1.append(["total:", total_sum])

wb.save("C:\\Users\\wy\\Desktop\\results.xlsx")

end_time = time.time()
execution_time = end_time - start_time
print(f"程序运行时间: {execution_time:.2f} 秒")
